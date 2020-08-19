from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment
import numpy as np
import pandas as pd
import requests
import json

class COVID19:
	def __init__(self):
		self.date = datetime.today().strftime('%Y/%m/%d %H:%M:%S')
		self.workbook = Workbook()
		self.alignment = Alignment(vertical='center',horizontal='center')
		self.country_name1 = json.load(open('dt1.json'))
		self.country_name2 = json.load(open('dt2.json'))
		self.data1 = self.worldometers()
		self.data2 = self.ecdc()
		self.data3 = self.nCov2019()
		self.summary = self.Summarize_Data()
		self.USA_COVID()
		self.Output_Summary()
		self.workbook.save('COVID19.xlsx')

	def parse_data(self,url):
		resp = requests.get(url)
		soup = BeautifulSoup(resp.text,'lxml')
		return soup

	def write_in_excel(self,data,worksheet):
		data['資料時間'] = self.date
		data.sort_values('確診數',ascending=False,inplace=True)
		for i,col in enumerate(data.columns):
			worksheet[chr(66+i)+'1'] = col
		for i,index in enumerate(data.index):
			worksheet['A'+str(2+i)] = index
		for row,record in enumerate(data.values):
			for col,val in enumerate(record):
				worksheet[chr(66+col)+str(2+row)] = val

	def worldometers(self):	# 資料來源 worldometers
		url = 'https://www.worldometers.info/coronavirus/'
		soup = self.parse_data(url)
		table = soup.find('table',attrs={'id':'main_table_countries_today'}).find('tbody')
		col = ['國家','確診數','新增案例數','死亡數','新增死亡數','治癒數','未治癒數','重症數']
		record = [[j.text.strip() for j in i.find_all('td')[1:9]] for i in table.find_all('tr')[8:-1]]
		data = pd.DataFrame(record,columns=col).set_index('國家')
		# 數值處理		
		func = lambda x:''.join(x.split('+')[-1].split(','))
		data = data.applymap(func)
		data.replace({'':np.nan,'N/A':np.nan},inplace=True)
		data.fillna(0,inplace=True)
		data = data.applymap(int)
		# 寫入 worksheet
		ws = self.workbook.create_sheet('worldometers')
		self.write_in_excel(data,ws)
		return data

	def ecdc(self):
		url = 'https://opendata.ecdc.europa.eu/covid19/casedistribution/json/'
		resp = requests.get(url).json()['records']
		raw_data = pd.DataFrame(resp).set_index('continentExp')
		stat_col = ['cases','deaths']
		for col in stat_col:
			raw_data[col] = raw_data[col].apply(int)
		df = raw_data.groupby('countriesAndTerritories').sum()[stat_col]
		country_info = raw_data['countriesAndTerritories'].drop_duplicates()
		data = pd.merge(df,country_info,how='left',left_index=True,right_on='countriesAndTerritories')
		data.columns = ['確診數','死亡數','國家']

		data.insert(0,'洲',data.index)
		data.set_index('國家',inplace=True)
		data.sort_values('確診數',ascending=False,inplace=True)
		data.index = [country.replace('_',' ') for country in data.index]
		data.index = [self.country_name1[country] if country in self.country_name1.keys() else country for country in data.index]
		ws = self.workbook.create_sheet('ecdc')
		self.write_in_excel(data,ws)
		return data
		
	def nCov2019(self):
		url = 'https://ncov2019.live/data'
		soup = self.parse_data(url)
		table = soup.find('table',attrs={'id':'sortable_table_world'})
		record = [list(record.stripped_strings)[1:] for record in table.find_all('tr')]
		data = pd.DataFrame(record[2:]).drop([3,4,7,8],axis=1)
		data.columns = ['國家','確診數','新增案例數','死亡數','新增死亡數','未治癒數','治癒數']
		data.replace('Unknown',np.nan,inplace=True)
		data.fillna(0,inplace=True)
		data.set_index('國家',inplace=True)
		data.index = [self.country_name2[country] if country in self.country_name2.keys() else country for country in data.index]
		data = data.applymap(str)
		data = data.applymap(lambda x: ''.join(x.split(',')))
		data = data.applymap(int)
		# 寫入 worksheet
		ws = self.workbook.create_sheet('nCov2019')
		self.write_in_excel(data,ws)
		return data

	def Summarize_Data(self):
		# 彙整各資料源的確診數(summary1),死亡數(summary2),治癒數(summary3)
		summary1 = self.data2[['洲','確診數']]
		summary1 = pd.merge(summary1,self.data1['確診數'],left_index=True,right_index=True,suffixes=('_ecdc','_worldometers'))
		summary1 = pd.merge(summary1,self.data3['確診數'],left_index=True,right_index=True)
		
		summary2 = self.data2['死亡數']
		summary2 = pd.merge(summary2,self.data1['死亡數'],left_index=True,right_index=True,suffixes=('_ecdc','_worldometers'))
		summary2 = pd.merge(summary2,self.data3['死亡數'],left_index=True,right_index=True)

		summary3 = self.data1['治癒數']
		summary3 = pd.merge(summary3,self.data3['治癒數'],left_index=True,right_index=True,suffixes=('_worldometers',''))

		summary = pd.merge(summary1,summary2,left_index=True,right_index=True)
		summary = pd.merge(summary,summary3,left_index=True,right_index=True)
		# summary.drop('China',inplace=True)
		summary.sort_values('洲',inplace=True)

		# 從亞洲國家中區分出中東地區
		Eastern_Asia = ('Israel','UAE','Egypt','Iran','Lebanon','Cyprus','Palestine','Iraq','Kuwait','Oman','Bahrain','Azerbaijan'
						,'Qatar','Saudi Arabia','Jordan','Turkey','Uzbekistan','Kyrgyzstan','Syria')
		for country in summary.index:
			if country in Eastern_Asia:
				summary.loc[country,'洲'] = 'Eastern Mediterranean'
		summary.sort_values('洲',inplace=True)

		col = [(item,resource) for item in ('確診數','死亡數','治癒數') for resource in ('ecdc','worldometers','nCov2019')]
		col.remove(('治癒數','ecdc'))
		col.insert(0,('洲','Continent'))
		summary.columns = pd.Index(col)
		return summary

	def Output_Summary(self):
		ws = self.workbook.create_sheet('Summary')
		for i in ('C1:E1','F1:H1','I1:J1'):
			ws.merge_cells(i)
		column = {'B1':'洲','C1':'確診數','F1':'死亡數','I1':'治癒數'}
		for cell,col in column.items():
			ws[cell] = col
			ws[cell].alignment = self.alignment

		column = [resource for _ in range(2) for resource in ('ecdc','worldometers','nCov2019')]
		column += ['worldometers','eCov2019']
		column.insert(0,'Continent')
		for col,name in enumerate(column):
			ws[chr(66+col)+'2'] = name
		for row,idx in enumerate(self.summary.index):
			ws['A'+str(3+row)] = idx
		for row,record in enumerate(self.summary.values):
			for col,val in enumerate(record):
				ws[chr(66+col)+str(3+row)] = val

	def USA_COVID(self):
		# 資料源一: nCov2019
		url1 = 'https://ncov2019.live/data'
		soup = self.parse_data(url1)
		table = soup.find('table',attrs={'id':'sortable_table_unitedstates'})
		record = [list(record.stripped_strings)[1:] for record in table.find_all('tr')]
		data1 = pd.DataFrame(record[2:])[[0,1,5,10]]
		data1 = data1.applymap(lambda x: x.replace(',',''))
		data1.columns = ['地區','確診數','死亡數','治癒數']
		data1['確診數'] = data1['確診數'].apply(int)
		data1.sort_values('確診數',ascending=False,inplace=True)

		# 資料源二: worldmeters
		url2 = 'https://www.worldometers.info/coronavirus/country/us/'
		soup = self.parse_data(url2)
		table = soup.find('table',attrs={'id':'usa_table_countries_today'})
		record = [[val.text for val in record.find_all('td')] for record in table.find_all('tr')]
		data2 = pd.DataFrame(record[2:-1])[[1,2,4,6]]
		data2 = data2.applymap(lambda x: x.replace('\n',''))
		for i in (2,4,6):
			data2[i] = data2[i].apply(lambda x: x.replace(',',''))
			data2[i] = data2[i].replace('N/A ',np.nan)
			data2[i] = data2[i].replace('',np.nan)
		data2.dropna(inplace=True)
		for i in (2,4,6):
			data2[i] = data2[i].apply(int)
		data2.columns = ['地區','確診數','死亡數','未治癒數']
		data2['治癒數'] = data2['確診數']-data2['死亡數']-data2['未治癒數']
		del data2['未治癒數']
		data2.sort_values('確診數',ascending=False,inplace=True)
		
		# 寫入 worksheet
		ws = self.workbook.create_sheet('USA_stats')
		dt = {'B1:E1':'nCov2019','G1:J1':'worldmeters'}
		for area,resource in dt.items():
			ws.merge_cells(area)
			locate = area.split(':')[0]
			ws[locate] = resource
			ws[locate].alignment = self.alignment
		for i in range(15):
			ws['A'+str(3+i)] = i+1
			ws['A'+str(3+i)].alignment = self.alignment
		for i,data in enumerate([data1,data2]):
			data = data.head(15)
			for num,col_name in enumerate(['地區','確診數','死亡數','治癒數']):
				ws[chr(66+num+i*5)+'2'] = col_name
			for row,record in enumerate(data.values):
				for col,val in enumerate(record):
					ws[chr(66+col+i*5)+str(3+row)] = val

		sheet = self.workbook['Sheet']
		self.workbook.remove(sheet)


if __name__ == "__main__":
	data = COVID19()
	data1 = data.data1
	data2 = data.data2
	data3 = data.data3
	country1 = set(data.data1.index)
	country2 = set(data.data2.index)
	country3 = set(data.data3.index)
	result = country1.intersection(country2,country3)
	summary = data.summary